import openpyxl as xl

# Book
wb = xl.load_workbook('test-data/test-data.xlsx')

# Sheet
ws = wb['Test3']

def vlookup(key, keyColumnAlphabet, resultColumnAlphabet):
    """keyColumnAlphabet列から key を探し、見つけた行の resultColumnAlphabet列のセルを返します
    Parameters
    ----------
    key : str
        探す値
    keyColumnAlphabet : str
        探す列
    resultColumnAlphabet : str
        欲しい値がある列
    """
    for rowNumber in range(2,10):
        id = ws[f'{keyColumnAlphabet}{rowNumber}'].value
        print(f'id={id}')
        if id == key:
            return ws[f'{resultColumnAlphabet}{rowNumber}']
        elif id is None or id == '':
            return None

    return None

result = vlookup(3, 'A', 'B')
print(f'found={result.value}')

from openpyxl import Workbook
wb = Workbook()

ws = wb.active

# insert at the end (default)
ws1 = wb.create_sheet("Mysheet")

# insert at first position
ws2 = wb.create_sheet("Mysheet", 0)

# insert at the penultimate position
ws3 = wb.create_sheet("Mysheet", -1)

ws.title = "New Title"

ws.sheet_properties.tabColor = "1072BA"

ws3 = wb["New Title"]

print(wb.sheetnames)

for sheet in wb:
    print(sheet.title)

source = wb.active
target = wb.copy_worksheet(source)

# Saving to a file
wb.save('./output/tutorial-1-create-a-workbook.xlsx')
